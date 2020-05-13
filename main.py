import urllib.request
import json
import time
from openpyxl import load_workbook

def get_cities_id(): 
	letters_rus = [
		'%D0%90',
		'%D0%91',
		'%D0%92',
		'%D0%93',
		'%D0%94',
		'%D0%95',
		'%D0%81',
		'%D0%96',
		'%D0%97',
		'%D0%98',
		'%D0%99',
		'%D0%9A',
		'%D0%9B',
		'%D0%9C',
		'%D0%9D',
		'%D0%9E',
		'%D0%9F',
		'%D0%A0',
		'%D0%A1',
		'%D0%A2',
		'%D0%A3',
		'%D0%A4',
		'%D0%A5',
		'%D0%A6',
		'%D0%A7',
		'%D0%A8',
		'%D0%A9',
		'%D0%AA',
		'%D0%AB',
		'%D0%AC',
		'%D0%AD',
		'%D0%AE',
		'%D0%AF'
	]

	cities_id = []
	for letter in letters_rus:
		full_api_url = 'https://www.homecredit.ru/api/v1/office/town/filter/?name=' + letter

		with urllib.request.urlopen(full_api_url) as response:
			JSON_response = response.read()

		JSON = json.loads(JSON_response)

		for item in JSON['data']:
			city_id = item['id']
			city_name = item['name']

			cities_id.append({
				'city_id': city_id,
				'city_name': city_name
		    })

	return cities_id

def get_branch_data(cities_id):
	branches_data = []
	for item in cities_id:
		JSON_branches_api = 'https://www.homecredit.ru/api/office/filter/?town=' + str(item['city_id'])

		with urllib.request.urlopen(JSON_branches_api) as response:
			JSON_response = response.read()

		JSON = json.loads(JSON_response)
		for branch in JSON['data']['data']:
			branch_name = branch['name']
			address = branch['address']
			city = branch['town']['name']
			lat = branch['ll']['lat']
			lng = branch['ll']['lng']

			branches_data.append({
				'branch_name': branch_name,
				'address': address,
				'city': city,
				'lat': lat,
				'lng': lng,
		    })
			print('Collecting:', branch_name, ';', address, ';', city, ';', lat, ';', lng)

		time.sleep(2)

	return branches_data

def write_xlsx(branches_data):
	workbook = load_workbook('32. OOO Home Credit and Finance Bank Russia.xlsx')
	worksheet = workbook[workbook.sheetnames[0]]
	cell_value = '2'

	for item in branches_data:
		branch_name = item['branch_name']
		lat = item['lat']
		lng = item['lng']
		city = item['city']
		address = item['address']

		cell_value = str(cell_value)
		worksheet['B' + cell_value] = 'OOO Home Credit and Finance Bank'
		worksheet['C' + cell_value] = branch_name
		worksheet['D' + cell_value] = address
		worksheet['G' + cell_value] = 'Russia'
		worksheet['H' + cell_value] = 'RU'
		worksheet['J' + cell_value] = city
		worksheet['M' + cell_value] = lat
		worksheet['N' + cell_value] = lng
		worksheet['O' + cell_value] = 'Address'
		worksheet['R' + cell_value] = 'Bank website'
		cell_value = int(cell_value) + 1

	workbook.save('32. OOO Home Credit and Finance Bank Russia.xlsx')

def main():
	cities_id = get_cities_id()
	branches_data = get_branch_data(cities_id)
	write_xlsx(branches_data)

if __name__ == '__main__':
    main()